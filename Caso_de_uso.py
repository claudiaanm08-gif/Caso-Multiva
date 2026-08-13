# -*- coding: utf-8 -*-
"""
Caso de uso - Especialista GI
Script único para los 6 pasos del caso.

DISEÑO (importante):
  - Todo lo que es un CÁLCULO (totales, %, tablas, deduplicación, detección
    de outliers, formatos, etc.) se resuelve 100% en Python/pandas. Es
    determinista, reproducible y no depende de ningún modelo.
  - Todo lo que es REDACCIÓN/ANÁLISIS EN LENGUAJE NATURAL (explicar de dónde
    sale cada diferencia, señalar hallazgos, criticar el borrador y
    reescribirlo, recomendar qué automatizar) se genera con un modelo de
    lenguaje OPEN SOURCE que corre LOCAL vía Ollama (https://ollama.com),
    sin API keys y sin mandar datos a internet.
  - Si Ollama no está instalado/corriendo, el script NO inventa el texto:
    dejo la sección marcada como pendiente, con instrucciones de cómo
    activarla. Las cifras y tablas se generan siempre, con o sin IA.

Cómo activar la parte generativa (opcional):
    1. Instala Ollama: https://ollama.com/download
    2. Descarga un modelo open source, por ejemplo:
         ollama pull llama3.1        (Meta, 8B, recomendado)
         ollama pull mistral         (Mistral 7B, más ligero)
       (basta con uno de los dos)
    3. Dale una sola vez ejecución en segundo plano y corre este script; se
       detecta automáticamente en http://localhost:11434.

Uso:
    Coloca este .py en la MISMA carpeta que 'Caso_Especialista_GI_CANDIDATO.xlsx'
    y ejecuta:  python Caso_de_uso.py

Requisitos:
    pip install pandas openpyxl
    (Ollama es opcional; sin él el script corre igual, solo sin narrativa)
"""

import json
import re
import difflib
import urllib.request
import urllib.error
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment

# Config
CARPETA = Path(__file__).resolve().parent
ARCHIVO_ENTRADA = CARPETA / "Caso_Especialista_GI_CANDIDATO.xlsx"
ARCHIVO_SALIDA = CARPETA / "Caso_Especialista_GI_CANDIDATO_LIMPIO.xlsx"
ARCHIVO_REGLAS = CARPETA / "reglas_aprendidas.json"
HOJA_DATOS = "Datos"

OLLAMA_URL = "http://localhost:11434"
MODELO_IA = "llama3.1"   

if not ARCHIVO_ENTRADA.exists():
    raise FileNotFoundError(
        f"No se encontró '{ARCHIVO_ENTRADA.name}' en {CARPETA}. "
        f"Verifica que el .py y el .xlsx estén en la misma carpeta."
    )


# CAPA GENERATIVA: llama con modelo open source LOCAL vía Ollama
_ia_disponible = None  # se detecta una sola vez por corrida


def ia_disponible():
    global _ia_disponible
    if _ia_disponible is not None:
        return _ia_disponible
    try:
        urllib.request.urlopen(f"{OLLAMA_URL}/api/tags", timeout=2)
        _ia_disponible = True
    except Exception:
        _ia_disponible = False
    return _ia_disponible


def generar_con_ia(prompt, max_tokens=600):
    """Genera texto con un modelo open source local (Ollama). Devuelve None
    si el modelo no está disponible o falla la llamada; el script NUNCA
    rellena ese None con texto inventado por el propio script."""
    if not ia_disponible():
        return None
    payload = {
        "model": MODELO_IA,
        "prompt": prompt,
        "stream": False,
        "options": {"num_predict": max_tokens, "temperature": 0.2},
    }
    try:
        req = urllib.request.Request(
            f"{OLLAMA_URL}/api/generate",
            data=json.dumps(payload).encode("utf-8"),
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=120) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            texto = data.get("response", "").strip()
            return texto if texto else None
    except (urllib.error.URLError, TimeoutError, Exception) as e:
        print(f"  [IA] No se pudo generar con Ollama/{MODELO_IA} ({e}).")
        return None


def parsear_numero_pct(cadena):
    """Convierte un porcentaje citado en texto a float, reconociendo tanto
    notación en inglés (89.6) como en español (89,6 con coma decimal) o
    europea mixta (1.234,5). El regex que llama a esta función solo captura
    porcentajes, que en este caso nunca superan ~10,625 — así que una coma
    seguida de 1-2 dígitos es casi siempre decimal, no miles."""
    s = cadena.replace("%", "").strip()
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")   # formato europeo: 1.234,5
        else:
            s = s.replace(",", "")                       # formato inglés: 1,234.5
    elif "," in s:
        partes = s.split(",")
        if len(partes) == 2 and 1 <= len(partes[1]) <= 2:
            s = s.replace(",", ".")                       # coma decimal: 89,6
        else:
            s = s.replace(",", "")                        # coma de miles: 10,625
    return float(s)


def verificar_cifras(texto, valores_pct_validos=None, valores_monto_validos=None,
                      tolerancia_pct=0.5, tolerancia_monto_rel=0.01):
    """Los modelos de lenguaje pueden citar mal un número aunque se les haya
    dado el dato exacto (es justo el riesgo que ilustra el Paso 4 del caso).
    Esta función NO corrige el texto: extrae los porcentajes Y los montos en
    pesos que menciona, y señala cuáles no coinciden con ninguna cifra de
    las tablas ya calculadas (o, si aplica, con las cifras que el propio
    borrador original citaba, para no marcar como error el que la crítica
    cite el número incorrecto del borrador al señalarlo)."""
    if not texto:
        return []
    sospechosos = []

    # NOTA: 0% y 100% son puntos de referencia universales
    PCT_SIEMPRE_PERMITIDOS = [0.0, 100.0]

    if valores_pct_validos is not None:
        for cita in re.findall(r"-?\d[\d,]*\.?\d*\s?%", texto):
            try:
                num = parsear_numero_pct(cita)
            except ValueError:
                continue
            if abs(num) in PCT_SIEMPRE_PERMITIDOS:
                continue

            if not any(abs(abs(num) - abs(v)) <= tolerancia_pct for v in valores_pct_validos if v is not None):
                sospechosos.append(cita.strip())

    if valores_monto_validos is not None:
        # montos con separador de miles (al menos una coma), con o sin '$'/'MXN'
        for cita in re.findall(r"\$?\s?-?\d{1,3}(?:,\d{3})+(?:\.\d+)?\s?(?:MXN)?", texto):
            try:
                num = float(cita.replace("$", "").replace("MXN", "").replace(",", "").strip())
            except ValueError:
                continue
            tol = max(1000, abs(num) * tolerancia_monto_rel)
            if not any(abs(abs(num) - abs(v)) <= tol for v in valores_monto_validos if v is not None):
                sospechosos.append(cita.strip())

    return sospechosos


def nota_verificacion(texto, sospechosos):
    partes = []
    if sospechosos:
        partes.append(f"⚠️ Verificación automática: el texto generado por IA cita {sospechosos} que "
                       f"NO coincide con ninguna cifra de las tablas calculadas arriba. Los modelos de "
                       f"lenguaje pueden citar mal un dato aunque se les dé el valor exacto — revisar "
                       f"manualmente antes de publicar este texto.")
    else:
        partes.append("✅ Verificación automática: los porcentajes y montos citados coinciden con "
                       "las tablas calculadas.")
    if parece_truncado(texto):
        partes.append("⚠️ El texto parece cortado a media oración (se quedó sin espacio de "
                       "generación). Vuelve a correr el script o ajusta 'max_tokens' en el código.")
    return " ".join(partes)


def extraer_numeros_de_texto(texto):
    """Extrae porcentajes y montos de un texto (p. ej. el borrador original)
    para poder permitir que la crítica los cite sin marcarlos como error."""
    if not texto:
        return [], []
    pcts = [parsear_numero_pct(c) for c in re.findall(r"-?\d[\d,]*\.?\d*\s?%", texto)]
    montos = [float(c.replace("$", "").replace("MXN", "").replace(",", "").strip())
              for c in re.findall(r"\$?\s?-?\d{1,3}(?:,\d{3})+(?:\.\d+)?\s?(?:MXN)?", texto)]
    return pcts, montos




def parece_truncado(texto):
    """Señal simple de que el modelo se quedó sin espacio (tokens) a media
    idea: el texto no termina en un cierre de oración típico."""
    if not texto:
        return False
    return texto.strip()[-1] not in ".!?\"'）)”»"


PENDIENTE_IA = (
    "⚠️ PENDIENTE — esta sección requiere el modelo de IA local (Ollama) para "
    "redactarse automáticamente y no se encontró corriendo en "
    f"{OLLAMA_URL}. Instala Ollama, corre 'ollama pull {MODELO_IA}' y vuelve "
    "a ejecutar el script. Mientras tanto, las cifras que sustentan esta "
    "sección sí quedan calculadas abajo."
)

print(f"Estado del modelo de IA local (Ollama): "
      f"{'DISPONIBLE (' + MODELO_IA + ')' if ia_disponible() else 'NO DISPONIBLE (se deja lo narrativo pendiente)'}")



def escribir_bloque(ws, fila_inicio, titulo, dataframe):
    ws.cell(row=fila_inicio, column=1, value=titulo).font = Font(bold=True)
    fila = fila_inicio + 1
    for j, col in enumerate(dataframe.columns, start=1):
        ws.cell(row=fila, column=j, value=col).font = Font(bold=True)
    fila += 1
    for _, row in dataframe.iterrows():
        for j, col in enumerate(dataframe.columns, start=1):
            ws.cell(row=fila, column=j, value=row[col])
        fila += 1
    return fila + 1


def escribir_texto(ws, fila_inicio, titulo, texto):
    ws.cell(row=fila_inicio, column=1, value=titulo).font = Font(bold=True)
    fila_inicio += 1
    ws.cell(row=fila_inicio, column=1, value=texto)
    ws.cell(row=fila_inicio, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    return fila_inicio + 2


# PASO 1: VALIDAR Y LIMPIAR
print("\nPASO 1: VALIDAR Y LIMPIAR\n")

df = pd.read_excel(ARCHIVO_ENTRADA, sheet_name=HOJA_DATOS, dtype=str)
df.columns = [c.strip() for c in df.columns]
df_original = df.copy()

problemas_encontrados = []  # lista problemas detectados


def registrar_problema(msg):
    problemas_encontrados.append(msg)
    print(f"  • {msg}")


print(f"Filas leídas: {len(df)} | Columnas: {list(df.columns)}")

# Perfilado automático por columna (detecta patrones con regex) 
PATRONES = {
    "fecha_dmy": re.compile(r"^\d{2}/\d{2}/\d{4}$"),
    "fecha_ymd": re.compile(r"^\d{4}-\d{2}-\d{2}$"),
    "fecha_ymd_hora": re.compile(r"^\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}$"),  # Excel guardó la celda
    "fecha_dmony": re.compile(r"^\d{2}-[A-Za-z]{3}-\d{4}$"),                   # como fecha real, no texto
    "numero_con_comas": re.compile(r"^-?\d{1,3}(,\d{3})+$"),
    "numero_sin_comas": re.compile(r"^-?\d+$"),
}


def perfilar_columna(nombre_col, serie):
    total = len(serie)
    nulos = int(serie.isna().sum())
    valores = serie.dropna().astype(str)
    patrones = {n: int(valores.apply(lambda x: bool(rx.match(x.strip()))).sum())
                for n, rx in PATRONES.items() if valores.apply(lambda x: bool(rx.match(x.strip()))).sum() > 0}
    return {"columna": nombre_col, "nulos": nulos,
            "pct_nulos": round(100 * nulos / total, 1) if total else 0,
            "valores_unicos": int(valores.nunique()), "patrones": patrones}


perfil = [perfilar_columna(c, df[c]) for c in df.columns]

# Espacios en blanco 
cols_texto = ["ID_Operacion", "Región", "Sucursal", "Ejecutivo", "Segmento", "Producto"]
for col in cols_texto:
    antes = df[col].astype(str)
    despues = antes.str.strip().replace({"nan": np.nan, "": np.nan})
    n = int((antes != despues.astype(str)).sum())
    if n > 0:
        registrar_problema(f"Columna '{col}': {n} valor(es) con espacios en blanco extra.")
    df[col] = despues

# Fecha: formatos mezclados -> AAAA-MM-DD
formatos_fecha = {p["columna"]: p["patrones"] for p in perfil if p["columna"] == "Fecha"}["Fecha"]
n_formatos = sum(1 for k in formatos_fecha if k.startswith("fecha_"))
if n_formatos > 1:
    registrar_problema(f"Columna 'Fecha': {n_formatos} formatos distintos mezclados {formatos_fecha} "
                        f"-> se estandarizan a AAAA-MM-DD.")


def parsear_fecha(valor):
    if pd.isna(valor) or str(valor).strip() == "":
        return pd.NaT
    v = str(valor).strip()
    # NOTA: Excel a veces guarda la celda como fecha real (no texto); al leerla con dtype=str, pandas la vuelve "AAAA-MM-DD HH:MM:SS". Se cubre explícito para no caer en el parser genérico con dayfirst (ambiguo y con warning).
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return pd.to_datetime(v, format=fmt)
        except ValueError:
            continue
    return pd.to_datetime(v, errors="coerce")


df["Fecha"] = df["Fecha"].apply(parsear_fecha)

# Segmento: variantes de escritura -> clustering por similitud 
def normalizar_texto_base(t):
    return t.strip().lower()


def sugerir_reglas_categoricas(serie, umbral=0.72):
    conteo_raw = serie.dropna().value_counts().to_dict()
    normalizados = sorted(set(normalizar_texto_base(v) for v in conteo_raw))
    grupos, asignados = [], set()
    for v in normalizados:
        if v in asignados:
            continue
        grupo = {v}
        asignados.add(v)
        for w in normalizados:
            if w in asignados:
                continue
            ratio = difflib.SequenceMatcher(None, v, w).ratio()
            comparten_raiz = len(v) >= 5 and len(w) >= 5 and v[:5] == w[:5]
            if ratio >= umbral or comparten_raiz:
                grupo.add(w)
                asignados.add(w)
        grupos.append(grupo)
    reglas = {}
    for grupo in grupos:
        candidatos = [r for r in conteo_raw if normalizar_texto_base(r) in grupo]
        canon = max(candidatos, key=lambda r: conteo_raw[r]).strip().capitalize()
        for r in candidatos:
            reglas[r] = canon
    return reglas


def cargar_reglas_previas(path):
    if not path.exists():
        return {}
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def guardar_reglas(path, reglas):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(reglas, f, ensure_ascii=False, indent=2, sort_keys=True)


reglas_previas = cargar_reglas_previas(ARCHIVO_REGLAS)
valores_originales_segmento = sorted(df["Segmento"].dropna().unique())
reglas_segmento = {**sugerir_reglas_categoricas(df["Segmento"]), **reglas_previas.get("Segmento", {})}
registrar_problema(f"Columna 'Segmento': {len(valores_originales_segmento)} variantes de escritura "
                    f"{valores_originales_segmento} -> unificadas a {sorted(set(reglas_segmento.values()))}.")
df["Segmento"] = df["Segmento"].map(reglas_segmento)
reglas_previas["Segmento"] = reglas_segmento
guardar_reglas(ARCHIVO_REGLAS, reglas_previas)

# Monto_Colocado: texto con/sin comas -> numérico 
def limpiar_monto(v):
    if pd.isna(v):
        return np.nan
    v = str(v).replace(",", "").replace("$", "").strip()
    try:
        return float(v)
    except ValueError:
        return np.nan


formato_monto = df_original["Monto_Colocado"].astype(str).apply(lambda x: "," in x)
if not formato_monto.all() and formato_monto.any():
    registrar_problema(f"Columna 'Monto_Colocado': formato numérico inconsistente "
                        f"({formato_monto.sum()} valores con comas, {(~formato_monto).sum()} sin comas) "
                        f"-> convertida a numérico puro.")
df["Monto_Colocado"] = df["Monto_Colocado"].apply(limpiar_monto)

# Deduplicación inteligente (sobrevive el registro más completo)
dup_mask = df.duplicated(subset=["ID_Operacion"], keep=False)
if dup_mask.any():
    ids_dup = sorted(df.loc[dup_mask, "ID_Operacion"].unique())
    filas_supervivientes = []
    for _id, grupo in df[dup_mask].groupby("ID_Operacion"):
        g = grupo.copy()
        g["_completitud"] = g.notna().sum(axis=1)
        g = g.sort_values(by=["_completitud", "Fecha"], ascending=[False, False])
        filas_supervivientes.append(g.iloc[[0]].drop(columns=["_completitud"]))
    sobrevivientes = pd.concat(filas_supervivientes, ignore_index=True)
    df = pd.concat([df[~dup_mask], sobrevivientes], ignore_index=True)
    registrar_problema(f"ID_Operacion duplicado: {ids_dup} ({dup_mask.sum()} filas). "
                        f"Se conserva el registro más completo de cada grupo.")

# Valores faltantes 
faltantes = df.isna().sum()
faltantes = faltantes[faltantes > 0]
for col, n in faltantes.items():
    ids_afectados = df.loc[df[col].isna(), "ID_Operacion"].tolist()
    registrar_problema(f"Columna '{col}': {n} valor(es) faltante(s) en {ids_afectados}.")
if df["Ejecutivo"].isna().any():
    df["Ejecutivo"] = df["Ejecutivo"].fillna("Sin asignar")

# Valores atípicos (negativos y outliers estadísticos) 
negativos = df.loc[df["Monto_Colocado"] < 0, "ID_Operacion"].tolist()
if negativos:
    registrar_problema(f"Monto_Colocado NEGATIVO en {negativos} (no debería ocurrir en un reporte "
                        f"de colocación; se conserva marcado para validación, no se elimina).")

media = df["Monto_Colocado"].mean()
std = df["Monto_Colocado"].std()
outliers = [i for i in df.loc[(df["Monto_Colocado"] - media).abs() > 3 * std, "ID_Operacion"] if i not in negativos]
if outliers:
    registrar_problema(f"Monto_Colocado atípico (>3 desv. estándar de la media={media:,.0f}) en "
                        f"{outliers} (posible error de captura; se conserva marcado, no se elimina).")


def marcar_observacion(row):
    obs = []
    if row["Monto_Colocado"] < 0:
        obs.append("Monto negativo")
    if row["ID_Operacion"] in outliers:
        obs.append("Posible outlier")
    return "; ".join(obs)


df["Observacion"] = df.apply(marcar_observacion, axis=1)

# Validación final 
errores_validacion = []
if df["ID_Operacion"].duplicated().any():
    errores_validacion.append("Persisten IDs duplicados.")
if df["Fecha"].isna().any():
    errores_validacion.append(f"{df['Fecha'].isna().sum()} fecha(s) sin interpretar.")
segmentos_invalidos = set(df["Segmento"].dropna().unique()) - set(reglas_segmento.values())
if segmentos_invalidos:
    errores_validacion.append(f"Segmentos fuera de catálogo: {segmentos_invalidos}")
if df["Monto_Colocado"].isna().any():
    errores_validacion.append(f"{df['Monto_Colocado'].isna().sum()} Monto_Colocado no numérico.")

print("\n----- VALIDACIÓN FINAL -----")
if errores_validacion:
    for e in errores_validacion:
        print(f"  PENDIENTE: {e}")
else:
    print("  OK: sin duplicados, fechas válidas, segmentos en catálogo, montos numéricos.")

df["Fecha"] = pd.to_datetime(df["Fecha"]).dt.strftime("%Y-%m-%d")
df = df.sort_values("ID_Operacion").reset_index(drop=True)

print(f"\nFilas originales: {len(df_original)}  ->  Filas limpias: {len(df)}")
print(f"Total de problemas listados: {len(problemas_encontrados)}")

df.to_csv(CARPETA / "Datos_Limpios.csv", index=False)

perfil_df = pd.DataFrame([
    {"Columna": p["columna"], "Nulos": p["nulos"], "% Nulos": p["pct_nulos"],
     "Valores únicos": p["valores_unicos"], "Patrones detectados": str(p["patrones"])}
    for p in perfil
])
problemas_df = pd.DataFrame({"Problema detectado": problemas_encontrados})

with pd.ExcelWriter(ARCHIVO_SALIDA, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Datos_Limpios", index=False)
    problemas_df.to_excel(writer, sheet_name="Problemas_Detectados", index=False)
    perfil_df.to_excel(writer, sheet_name="Perfilado_Origen", index=False)




# PASO 2: CONCILIAR
print("\n\nPASO 2: CONCILIARn")

ctrl_raw = pd.read_excel(ARCHIVO_ENTRADA, sheet_name="Control", header=None)


def buscar_valor_por_etiqueta(df_ctrl, palabra_clave):
    for _, row in df_ctrl.iterrows():
        for j, val in enumerate(row):
            if isinstance(val, str) and palabra_clave.lower() in val.lower():
                for k in range(j + 1, len(row)):
                    c = row[k]
                    if isinstance(c, (int, float)) and not pd.isna(c):
                        return float(c)
    return None


total_control = buscar_valor_por_etiqueta(ctrl_raw, "Total de colocación")
meta_global = buscar_valor_por_etiqueta(ctrl_raw, "Meta global")

df_crudo = pd.read_excel(ARCHIVO_ENTRADA, sheet_name=HOJA_DATOS, dtype=str)
df_crudo["Monto_num"] = df_crudo["Monto_Colocado"].astype(str).str.replace(",", "", regex=False).astype(float)
total_crudo = df_crudo["Monto_num"].sum()

fila_dup = df_crudo[df_crudo.duplicated("ID_Operacion", keep=False)]
monto_dup = float(fila_dup["Monto_num"].iloc[0]) if len(fila_dup) else 0.0
total_tras_dedup = total_crudo - monto_dup

monto_outlier = float(df.loc[df["ID_Operacion"].isin(outliers), "Monto_Colocado"].sum())
monto_negativo = float(df.loc[df["ID_Operacion"].isin(negativos), "Monto_Colocado"].sum())
total_operaciones_normales = total_tras_dedup - monto_outlier - monto_negativo
diferencia_final = total_operaciones_normales - total_control

bridge_df = pd.DataFrame([
    {"Concepto": "Total original en 'Datos' (tal cual reportado)", "Monto (MXN)": total_crudo},
    {"Concepto": "(-) Registro(s) duplicado(s)", "Monto (MXN)": -monto_dup},
    {"Concepto": "= Subtotal tras deduplicar", "Monto (MXN)": total_tras_dedup},
    {"Concepto": "(-) Monto(s) atípico(s) (outlier)", "Monto (MXN)": -monto_outlier},
    {"Concepto": "(-) Monto(s) negativo(s)", "Monto (MXN)": -monto_negativo},
    {"Concepto": "= Total de operaciones 'normales'", "Monto (MXN)": total_operaciones_normales},
    {"Concepto": "Total de control (sistema)", "Monto (MXN)": total_control},
    {"Concepto": "DIFERENCIA sin explicar", "Monto (MXN)": diferencia_final},
])

print("----- PUENTE DE CONCILIACIÓN (cifras) -----")
print(bridge_df.to_string(index=False))

prompt_conciliacion = f"""Eres un analista financiero senior de un banco en México. Tienes estas cifras YA
CALCULADAS y VALIDADAS (no las recalcules, solo explícalas):

- Total original reportado en el sistema fuente, 56 filas, sin ninguna limpieza: {total_crudo:,.0f} MXN
- Un registro (ID_Operacion) estaba duplicado por {monto_dup:,.0f} MXN
- Tras quitar el duplicado, el subtotal es: {total_tras_dedup:,.0f} MXN
- Se detectó un monto atípico (outlier estadístico, >3 desviaciones estándar de la media) por
  {monto_outlier:,.0f} MXN en la(s) operación(es) {outliers}
- Se detectó (un) monto(s) negativo(s) por {monto_negativo:,.0f} MXN en la(s) operación(es) {negativos}
  (un monto negativo no debería existir en un reporte de colocación)
- Al aislar ambos, el total de operaciones "normales" queda en: {total_operaciones_normales:,.0f} MXN
- El total de control reportado por el sistema (ya validado por el área de sistemas) es:
  {total_control:,.0f} MXN
- Queda una diferencia residual sin explicar de: {diferencia_final:,.0f} MXN
  ({diferencia_final / total_control * 100:.1f}% del control)

Tarea: en español, en 4 párrafos cortos (uno por cada partida: duplicado, outlier, negativo,
diferencia residual), explica de dónde sale cada diferencia y por qué es relevante para el
negocio. No inventes cifras nuevas, usa solo las de arriba. Para la diferencia residual, sé
honesto: no se explica con la información disponible, sugiere 2-3 causas posibles a validar
con sistemas (por ejemplo: operaciones fuera del corte, otro canal no capturado, ajustes
manuales posteriores al reporte) sin afirmar cuál es la correcta."""

explicacion_conciliacion = generar_con_ia(prompt_conciliacion)

valores_pct_paso2 = [round(diferencia_final / total_control * 100, 1)]
valores_monto_paso2 = bridge_df["Monto (MXN)"].tolist()
sospechosos_paso2 = verificar_cifras(explicacion_conciliacion, valores_pct_paso2, valores_monto_paso2)

print("\n----- EXPLICACIÓN DE CADA DIFERENCIA -----")
print(explicacion_conciliacion if explicacion_conciliacion else PENDIENTE_IA)
if explicacion_conciliacion:
    print(f"\n  {nota_verificacion(explicacion_conciliacion, sospechosos_paso2)}")

wb = openpyxl.load_workbook(ARCHIVO_SALIDA)
ws = wb.create_sheet("Conciliacion")
fila = escribir_bloque(ws, 1, "Puente de conciliación", bridge_df)
fila = escribir_texto(ws, fila, "Explicación de cada diferencia (generado por IA local)",
                       explicacion_conciliacion or PENDIENTE_IA)
if explicacion_conciliacion:
    fila = escribir_texto(ws, fila, "Verificación automática", nota_verificacion(explicacion_conciliacion, sospechosos_paso2))
ws.column_dimensions["A"].width = 70
ws.column_dimensions["B"].width = 18
wb.save(ARCHIVO_SALIDA)





# PASO 3: ANALIZAR

print("\n\nPASO 3: ANALIZAR n")

metas = pd.read_excel(ARCHIVO_ENTRADA, sheet_name="Metas")
metas = metas[metas["Ejecutivo"].astype(str).str.upper() != "TOTAL"].copy()
metas["Meta_Semanal"] = pd.to_numeric(metas["Meta_Semanal"], errors="coerce")


ids_atipicos = set(outliers) | set(negativos)
df_ajustado = df[~df["ID_Operacion"].isin(ids_atipicos)]

por_ejec = df.groupby("Ejecutivo")["Monto_Colocado"].sum()
por_ejec_adj = df_ajustado.groupby("Ejecutivo")["Monto_Colocado"].sum()
cumpl_ejec = metas.copy()
cumpl_ejec["Colocado"] = cumpl_ejec["Ejecutivo"].map(por_ejec).fillna(0)
cumpl_ejec["Colocado (ajustado)"] = cumpl_ejec["Ejecutivo"].map(por_ejec_adj).fillna(0)
cumpl_ejec["% Cumplimiento"] = (cumpl_ejec["Colocado"] / cumpl_ejec["Meta_Semanal"] * 100).round(1)
cumpl_ejec["% Cumplimiento (ajustado)"] = (
    cumpl_ejec["Colocado (ajustado)"] / cumpl_ejec["Meta_Semanal"] * 100).round(1)
cumpl_ejec = cumpl_ejec.sort_values("% Cumplimiento (ajustado)", ascending=False).reset_index(drop=True)

por_region = df.groupby("Región")["Monto_Colocado"].sum()
por_region_adj = df_ajustado.groupby("Región")["Monto_Colocado"].sum()
meta_region = metas.groupby("Región")["Meta_Semanal"].sum()
cumpl_region = pd.DataFrame({"Meta_Semanal": meta_region, "Colocado": por_region,
                              "Colocado (ajustado)": por_region_adj}).fillna(0)
cumpl_region["% Cumplimiento"] = (cumpl_region["Colocado"] / cumpl_region["Meta_Semanal"] * 100).round(1)
cumpl_region["% Cumplimiento (ajustado)"] = (
    cumpl_region["Colocado (ajustado)"] / cumpl_region["Meta_Semanal"] * 100).round(1)
cumpl_region = cumpl_region.reset_index().rename(columns={"index": "Región"})

por_segmento = df.groupby("Segmento")["Monto_Colocado"].sum().sort_values(ascending=False)
por_segmento_adj = df_ajustado.groupby("Segmento")["Monto_Colocado"].sum()
dist_segmento = por_segmento.reset_index()
dist_segmento["% del total"] = (dist_segmento["Monto_Colocado"] / dist_segmento["Monto_Colocado"].sum() * 100).round(1)
dist_segmento["Monto (ajustado)"] = dist_segmento["Segmento"].map(por_segmento_adj).fillna(0)
dist_segmento["% del total (ajustado)"] = (
    dist_segmento["Monto (ajustado)"] / dist_segmento["Monto (ajustado)"].sum() * 100).round(1)

cols_ejec = ["Ejecutivo", "Región", "Meta_Semanal", "Colocado", "% Cumplimiento",
             "Colocado (ajustado)", "% Cumplimiento (ajustado)"]
cols_seg = ["Segmento", "Monto_Colocado", "% del total", "Monto (ajustado)", "% del total (ajustado)"]

print("----- % CUMPLIMIENTO POR EJECUTIVO (con atípicos / ajustado) -----")
print(cumpl_ejec[cols_ejec].to_string(index=False))
print("\n----- % CUMPLIMIENTO POR REGIÓN (con atípicos / ajustado) -----")
print(cumpl_region.to_string(index=False))
print("\n----- DISTRIBUCIÓN POR SEGMENTO (con atípicos / ajustado) -----")
print(dist_segmento[cols_seg].to_string(index=False))

prompt_hallazgos = f"""Eres un analista de datos de un banco en México. Tienes estas tablas YA
CALCULADAS (no las recalcules). Cada tabla trae dos versiones: "con atípicos" (los datos limpios
tal cual, incluyendo las operaciones {sorted(ids_atipicos)} marcadas como atípicas en un paso
previo) y "ajustado" (excluyendo esas operaciones atípicas).

% Cumplimiento por ejecutivo:
{cumpl_ejec[cols_ejec].to_string(index=False)}

% Cumplimiento por región:
{cumpl_region.to_string(index=False)}

Distribución de colocación por segmento:
{dist_segmento[cols_seg].to_string(index=False)}

Tarea: en español, entrega EXACTAMENTE 2 hallazgos relevantes para negocio (no más, no menos),
cada uno en un párrafo corto. Deben ser hallazgos que un analista destacaría a su jefe, no una
simple repetición de las tablas — usa la comparación "con atípicos" vs. "ajustado" cuando sea
relevante. IMPORTANTE: cita los porcentajes y montos EXACTOS de las tablas de arriba, sin
redondear distinto ni recalcular — cópialos tal cual aparecen."""

hallazgos_texto = generar_con_ia(prompt_hallazgos)

valores_pct_paso3 = (
    cumpl_ejec["% Cumplimiento"].tolist() + cumpl_ejec["% Cumplimiento (ajustado)"].tolist()
    + cumpl_region["% Cumplimiento"].tolist() + cumpl_region["% Cumplimiento (ajustado)"].tolist()
    + dist_segmento["% del total"].tolist() + dist_segmento["% del total (ajustado)"].tolist()
)
valores_monto_paso3 = (
    cumpl_ejec["Colocado"].tolist() + cumpl_ejec["Colocado (ajustado)"].tolist()
    + cumpl_ejec["Meta_Semanal"].tolist()
    + cumpl_region["Colocado"].tolist() + cumpl_region["Colocado (ajustado)"].tolist()
    + dist_segmento["Monto_Colocado"].tolist() + dist_segmento["Monto (ajustado)"].tolist()
)
sospechosos_paso3 = verificar_cifras(hallazgos_texto, valores_pct_paso3, valores_monto_paso3)

print("\n----- 2 HALLAZGOS RELEVANTES -----")
print(hallazgos_texto if hallazgos_texto else PENDIENTE_IA)
if hallazgos_texto:
    print(f"\n  {nota_verificacion(hallazgos_texto, sospechosos_paso3)}")

wb = openpyxl.load_workbook(ARCHIVO_SALIDA)
ws = wb.create_sheet("Analisis")
fila = escribir_bloque(ws, 1, "% Cumplimiento por Ejecutivo (con atípicos / ajustado)",
                        cumpl_ejec[cols_ejec])
fila = escribir_bloque(ws, fila, "% Cumplimiento por Región (con atípicos / ajustado)", cumpl_region)
fila = escribir_bloque(ws, fila, "Distribución por Segmento (con atípicos / ajustado)",
                        dist_segmento[cols_seg])
fila = escribir_texto(ws, fila, "2 hallazgos relevantes (generado por IA local)",
                       hallazgos_texto or PENDIENTE_IA)
if hallazgos_texto:
    fila = escribir_texto(ws, fila, "Verificación automática", nota_verificacion(hallazgos_texto, sospechosos_paso3))
ws.column_dimensions["A"].width = 22
for col in "BCDEFG":
    ws.column_dimensions[col].width = 20
wb.save(ARCHIVO_SALIDA)






print(f"\n\nArchivo final: {ARCHIVO_SALIDA}")
print(f"CSV de datos limpios: {CARPETA / 'Datos_Limpios.csv'}")
print(f"Reglas aprendidas: {ARCHIVO_REGLAS}")



